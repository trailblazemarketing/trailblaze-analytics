"""Illinois Gaming Board — sports wagering monthly reports.

IGB publishes monthly sports-wagering reports as XLSX on:
    https://www.igb.illinois.gov/SportsReports.aspx

Each workbook has a state-total sheet with handle and adjusted gross receipts
broken down by master licensee. We read the total row to capture state-level
handle and revenue.
"""

from __future__ import annotations

import calendar
import io
import re
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from trailblaze.scrapers.base import RegulatorScraper, ScrapedMetric
from trailblaze.scrapers.common import http_client
from trailblaze.scrapers.regulators._pdf import to_decimal

_MONTH_NAMES = {calendar.month_name[i].lower(): i for i in range(1, 13)}
_MONTH_ABBR = {calendar.month_abbr[i].lower(): i for i in range(1, 13)}


def _month_from_text(text: str) -> tuple[int, int] | None:
    t = text.lower()
    year_match = re.search(r"(20\d{2})", t)
    if not year_match:
        return None
    year = int(year_match.group(1))
    for name, num in _MONTH_NAMES.items():
        if name in t:
            return year, num
    for name, num in _MONTH_ABBR.items():
        if re.search(rf"\b{name}\b", t):
            return year, num
    return None


class IllinoisGamingBoardScraper(RegulatorScraper):
    name = "IL IGB"
    market_slug = "us-illinois"
    base_url = "https://www.igb.illinois.gov/SportsReports.aspx"
    # Old URL redirects to /sports-wagering which is a different page layout.
    # Needs index rewrite against the new IGB sports-wagering landing page.
    scraper_status = "broken_needs_research"

    def __init__(self, session, months: int = 6) -> None:
        super().__init__(session)
        self.months = months

    def scrape(self) -> list[ScrapedMetric]:
        records: list[ScrapedMetric] = []
        with http_client() as client:
            try:
                html = client.get(self.base_url).text
            except Exception as exc:
                self.log.warning("IL IGB: fetch index failed (%s)", exc)
                return records

            soup = BeautifulSoup(html, "lxml")
            candidates: list[tuple[int, int, str]] = []
            for a in soup.find_all("a", href=True):
                href = a["href"]
                text = a.get_text(" ", strip=True)
                lower = href.lower()
                if not (lower.endswith(".xlsx") or lower.endswith(".xls") or lower.endswith(".pdf")):
                    continue
                my = _month_from_text(text) or _month_from_text(href)
                if my is None:
                    continue
                candidates.append((my[0], my[1], urljoin(self.base_url, href)))

            candidates.sort(reverse=True)
            seen = 0
            for year, month, url in candidates:
                if seen >= self.months:
                    break
                seen += 1
                if url.lower().endswith((".xlsx", ".xls")):
                    try:
                        resp = client.get(url)
                        resp.raise_for_status()
                    except Exception as exc:
                        self.log.warning("IL IGB %04d-%02d: fetch failed (%s)",
                                         year, month, exc)
                        continue
                    records.extend(self._parse_xlsx(resp.content, year, month, url))
                # PDFs from IL are rarer; pattern-matching them is unreliable,
                # so we skip and let the XLSX path carry the load.

        return records

    def _parse_xlsx(self, content: bytes, year: int, month: int,
                    source_url: str) -> list[ScrapedMetric]:
        recs: list[ScrapedMetric] = []
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            self.log.warning("IL IGB %04d-%02d: cannot open xlsx (%s)", year, month, exc)
            return recs

        handle_total = None
        agr_total = None
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [c for c in row if c is not None]
                if not cells:
                    continue
                joined = " ".join(str(c).lower() for c in cells if isinstance(c, str))
                if "total" not in joined and "grand total" not in joined:
                    continue
                # Find numeric cells in the row — heuristic: largest two numbers
                # are handle (biggest) and AGR.
                numerics = [c for c in cells if isinstance(c, (int, float))]
                if len(numerics) < 2:
                    numerics = [to_decimal(str(c)) for c in cells]
                    numerics = [n for n in numerics if n is not None]
                if len(numerics) < 2:
                    continue
                numerics_sorted = sorted(numerics, reverse=True)
                if handle_total is None:
                    handle_total = numerics_sorted[0]
                if agr_total is None:
                    agr_total = numerics_sorted[1]
                if handle_total is not None and agr_total is not None:
                    break
            if handle_total is not None and agr_total is not None:
                break

        if handle_total is not None:
            recs.append(ScrapedMetric(
                metric_code="sportsbook_handle",
                period_year=year, period_month=month,
                value_numeric=handle_total, currency="USD",
                market_id=self.market_id,
                notes=f"IL IGB sports wagering monthly report {year}-{month:02d}",
                source_url=source_url,
            ))
        if agr_total is not None:
            recs.append(ScrapedMetric(
                metric_code="sportsbook_revenue",
                period_year=year, period_month=month,
                value_numeric=agr_total, currency="USD",
                market_id=self.market_id,
                notes=f"IL IGB sports wagering monthly report {year}-{month:02d}",
                source_url=source_url,
            ))
        return recs
